"""Universe seeding from the IDX stock-list xlsx."""

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # pyright: ignore[reportMissingTypeStubs]
from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.common.types import UniverseItem
from app.infrastructure.repositories.stock_repo import StockRepository

_MONTHS = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "Mei": 5,
    "Jun": 6,
    "Jul": 7,
    "Agt": 8,
    "Sep": 9,
    "Okt": 10,
    "Nov": 11,
    "Des": 12,
}


def parse_universe(path: Path) -> list[UniverseItem]:
    items: list[UniverseItem] = []
    for raw in _iter_rows(path):
        ticker = str(raw[1]).strip() if raw[1] is not None else ""
        if not ticker:
            continue
        items.append(
            UniverseItem(
                ticker=ticker,
                name=str(raw[2]).strip() if raw[2] is not None else "",
                listing_date=_parse_date(raw[3]),
                board=str(raw[5]).strip() if raw[5] is not None else "",
                shares_outstanding=_parse_shares(raw[4]),
            )
        )
    return items


async def seed_universe(session: AsyncSession, *, path: Path) -> int:
    items = parse_universe(path)
    return await StockRepository(session).upsert_stocks(items)


# openpyxl has no typed stubs; confine Any to this single boundary.
def _iter_rows(path: Path) -> list[tuple[Any, ...]]:
    workbook: Any = load_workbook(path, read_only=True, data_only=True)
    sheet: Any = workbook.active
    return [tuple(row) for row in sheet.iter_rows(min_row=2, values_only=True)]


def _parse_date(raw: object) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if not isinstance(raw, str):
        return None
    try:
        day, month, year = raw.strip().split(" ")
        return date(int(year), _MONTHS[month], int(day))
    except (ValueError, KeyError):
        return None


def _parse_shares(raw: object) -> int | None:
    if isinstance(raw, int):
        return raw
    if not isinstance(raw, str):
        return None
    digits = raw.strip().replace(".", "")
    return int(digits) if digits.isdigit() else None
